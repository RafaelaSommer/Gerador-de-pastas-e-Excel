import os
import ctypes
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk

import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


# ============================ CAMINHOS ============================ #
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PNG = os.path.join(BASE_DIR, "logo.png")
LOGO_ICO = os.path.join(BASE_DIR, "logo.ico")


def get_desktop_path():
    CSIDL_DESKTOPDIRECTORY = 0x10
    buf = ctypes.create_unicode_buffer(260)
    ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOPDIRECTORY, None, 0, buf)
    return buf.value


DESKTOP = get_desktop_path()


# ============================ TEMA ============================ #
BG = "#1e1e1e"
CARD = "#2b2b2b"
INPUT_BG = "#3a3a3a"
TEXT = "#f0f0f0"
BTN_BG = "#6C63FF"


# ============================ BOTÃO ============================ #
def styled_button(master, text, command):
    return tk.Button(
        master,
        text=text,
        command=command,
        bg=BTN_BG,
        fg="white",
        relief="flat",
        font=("Segoe UI", 11, "bold"),
        activebackground="#827CFF"
    )


# ============================ UTIL ============================ #
def centralizar_janela(janela, largura, altura):
    janela.update_idletasks()
    ws = janela.winfo_screenwidth()
    hs = janela.winfo_screenheight()
    x = (ws - largura) // 2
    y = (hs - altura) // 2
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


entradas_celulas = []


# ============================ POPUP SALVAR ============================ #
def popup_salvar_arquivo(parent):
    popup = tk.Toplevel(parent)
    popup.title("Salvar Excel")
    popup.configure(bg=CARD)
    popup.resizable(False, False)

    centralizar_janela(popup, 420, 200)

    resultado = {"nome": None}

    tk.Label(
        popup,
        text="Nome do arquivo Excel",
        bg=CARD,
        fg=TEXT,
        font=("Segoe UI", 12, "bold")
    ).pack(pady=(25, 10))

    entrada = tk.Entry(popup, font=("Segoe UI", 12), justify="center", width=30)
    entrada.pack(ipady=6)
    entrada.focus()

    def confirmar():
        if entrada.get().strip():
            resultado["nome"] = entrada.get().strip()
            popup.destroy()

    frame = tk.Frame(popup, bg=CARD)
    frame.pack(pady=25)

    styled_button(frame, "OK", confirmar).pack(side="left", padx=10)
    styled_button(frame, "Cancelar", popup.destroy).pack(side="left", padx=10)

    popup.transient(parent)
    popup.grab_set()
    parent.wait_window(popup)

    return resultado["nome"]


# ============================ GRADE ============================ #
def criar_grade():
    global entradas_celulas
    entradas_celulas.clear()

    for w in frame_grade_inner.winfo_children():
        w.destroy()

    try:
        cols = int(entry_colunas.get())
        rows = int(entry_linhas.get())
        if cols <= 0 or rows <= 0:
            raise ValueError
    except:
        messagebox.showerror("Erro", "Digite valores válidos.")
        return

    centralizar_janela(root, 1100, 650)

    for r in range(rows + 1):
        linha = []
        for c in range(cols):
            e = tk.Entry(
                frame_grade_inner,
                bg=INPUT_BG,
                fg=TEXT,
                font=("Segoe UI", 11, "bold" if r == 0 else "normal"),
                justify="center",
                relief="flat",
                insertbackground="white",
                width=18
            )
            e.grid(row=r, column=c, padx=10, pady=6, ipady=6)
            linha.append(e)
        entradas_celulas.append(linha)

    canvas_grade.configure(scrollregion=canvas_grade.bbox("all"))
    centralizar_grade()


# ============================ EXCEL ============================ #
def ajustar_excel(ws, dados):
    header_fill = PatternFill("solid", fgColor="404040")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, len(dados[0]) + 1):
        letra = get_column_letter(col)
        max_len = 0

        for row in range(1, len(dados) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

            if row == 1:
                cell.fill = header_fill
                cell.font = header_font

        ws.column_dimensions[letra].width = min(max_len + 5, 45)

    ws.freeze_panes = "A2"


# ============================ GERAR EXCEL ============================ #
def gerar_excel():
    if not entradas_celulas:
        messagebox.showerror("Erro", "Crie a grade primeiro.")
        return

    dados = []
    for r, linha in enumerate(entradas_celulas):
        valores = []
        for e in linha:
            if r == 0 and not e.get().strip():
                messagebox.showerror("Erro", "Preencha todos os cabeçalhos.")
                return
            valores.append(e.get().strip())
        dados.append(valores)

    nome = popup_salvar_arquivo(root)
    if not nome:
        return

    caminho = os.path.join(DESKTOP, f"{nome}.xlsx")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"

        for linha in dados:
            ws.append(linha)

        ajustar_excel(ws, dados)
        wb.save(caminho)

        messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{caminho}")

    except Exception as e:
        messagebox.showerror("Erro ao salvar", str(e))


# ============================ INTERFACE ============================ #
root = tk.Tk()
root.title("Gerador de Excel")
root.configure(bg=BG)
centralizar_janela(root, 520, 360)

if os.path.exists(LOGO_ICO):
    root.iconbitmap(LOGO_ICO)

if os.path.exists(LOGO_PNG):
    img = Image.open(LOGO_PNG).resize((72, 72))
    logo = ImageTk.PhotoImage(img)
    tk.Label(root, image=logo, bg=BG).pack(pady=8)

frame = tk.Frame(root, bg=CARD)
frame.pack(pady=10)

topo = tk.Frame(frame, bg=CARD)
topo.pack(padx=25, pady=15)

tk.Label(topo, text="Colunas", bg=CARD, fg=TEXT).grid(row=0, column=0, padx=15)
entry_colunas = tk.Entry(topo, bg=INPUT_BG, fg=TEXT, width=10, justify="center")
entry_colunas.grid(row=1, column=0, padx=15)

tk.Label(topo, text="Linhas", bg=CARD, fg=TEXT).grid(row=0, column=1, padx=15)
entry_linhas = tk.Entry(topo, bg=INPUT_BG, fg=TEXT, width=10, justify="center")
entry_linhas.grid(row=1, column=1, padx=15)

botoes = tk.Frame(frame, bg=CARD)
botoes.pack(pady=10)

styled_button(botoes, "Criar Grade", criar_grade).pack(side="left", padx=10)
styled_button(botoes, "Gerar Excel", gerar_excel).pack(side="left", padx=10)


# ============================ SCROLL ============================ #
container_grade = tk.Frame(root, bg=BG)
container_grade.pack(fill="both", expand=True, padx=20, pady=10)

canvas_grade = tk.Canvas(container_grade, bg=BG, highlightthickness=0)
canvas_grade.pack(side="left", fill="both", expand=True)

scroll = tk.Scrollbar(container_grade, orient="vertical", command=canvas_grade.yview)
scroll.pack(side="right", fill="y")

canvas_grade.configure(yscrollcommand=scroll.set)

frame_grade_inner = tk.Frame(canvas_grade, bg=BG)
canvas_window = canvas_grade.create_window((0, 0), window=frame_grade_inner, anchor="n")


def centralizar_grade(event=None):
    cw = canvas_grade.winfo_width()
    fw = frame_grade_inner.winfo_reqwidth()
    canvas_grade.coords(canvas_window, max((cw - fw) // 2, 0), 0)


frame_grade_inner.bind("<Configure>", lambda e: (
    canvas_grade.configure(scrollregion=canvas_grade.bbox("all")),
    centralizar_grade()
))

canvas_grade.bind("<Configure>", centralizar_grade)

root.mainloop()
