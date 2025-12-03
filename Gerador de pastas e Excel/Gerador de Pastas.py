import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog, font
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk  # Para carregar a logo

COLUMNS = {
    "Equipes": ["Equipe", "Realizado", "Não Realizado"],
    "Gerentes": ["Nome", "Em Atraso", "No Prazo"]
}

COLUMN_WIDTHS = {
    "Equipe": 30,
    "Realizado": 15,
    "Não Realizado": 15,
    "Nome": 30,
    "Em Atraso": 15,
    "No Prazo": 15
}

# =========================
# Tema Dark Mode (visual)
# =========================
BG = "#141414"
CARD = "#1e1e1e"
INPUT_BG = "#252525"
TEXT = "#EAEAEA"
ACCENT = "#11C620"
ACCENT_HOVER = "#11C620"
BTN_BG = "#000000"
BTN_FG = TEXT
TREE_BG = "#171717"
ROW_ALT = "#1b1b1b"

# Inicializa janela principal (aplicando fonte padrão)
default_font = ("Segoe UI", 10)
root = tk.Tk()
root.title("Gerador de Pastas")
root.configure(bg=BG)
root.option_add("*Font", default_font)

# Tentar ícone (não crítico)
try:
    caminho_ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.ico")
    if os.path.exists(caminho_ico):
        root.iconbitmap(caminho_ico)
except Exception:
    pass

# Centraliza janela
w, h = 1000, 700
ws, hs = root.winfo_screenwidth(), root.winfo_screenheight()
x, y = (ws - w)//2, (hs - h)//2
root.geometry(f"{w}x{h}+{x}+{y}")

# =========================
# Estilos ttk
# =========================
style = ttk.Style()
try:
    style.theme_use("clam")
except:
    pass

style.configure("TFrame", background=BG)
style.configure("Card.TFrame", background=CARD, relief="flat")
style.configure("TLabel", background=BG, foreground=TEXT)
style.configure("Card.TLabel", background=CARD, foreground=TEXT)
style.configure("TNotebook", background=BG, borderwidth=0)
style.configure("TNotebook.Tab", background=BTN_BG, foreground=TEXT, padding=(12, 8))
style.map("TNotebook.Tab", background=[("selected", ACCENT)], foreground=[("selected", "white")])

style.configure("TEntry", fieldbackground=INPUT_BG, background=INPUT_BG, foreground=TEXT, padding=5)
style.configure("TCombobox", fieldbackground=INPUT_BG, background=INPUT_BG, foreground=TEXT)

style.configure("Accent.TButton", background=ACCENT, foreground="white",
                font=(default_font[0], 10, "bold"))
style.map("Accent.TButton", background=[("active", ACCENT_HOVER)])

style.configure("TButton", background=BTN_BG, foreground=TEXT)

style.configure("Treeview",
                background=TREE_BG,
                fieldbackground=TREE_BG,
                foreground=TEXT,
                rowheight=26,
                bordercolor=BG,
                borderwidth=0)
style.configure("Treeview.Heading",
                background=ACCENT,
                foreground="white",
                font=(default_font[0], 10, "bold"))
style.map("Treeview", background=[("selected", ACCENT)])


# =========================
# Helper para botões tk
# =========================
def styled_button(parent, **kwargs):
    b = tk.Button(parent,
                  bg=ACCENT,
                  fg="white",
                  activebackground=ACCENT_HOVER,
                  activeforeground="white",
                  bd=0,
                  relief="flat",
                  highlightthickness=0,
                  font=(default_font[0], 10, "bold"),
                  **kwargs)
    return b


# =========================
# Aba 1: Gerador de Pastas
# =========================
class GeradorPastas:
    def __init__(self, parent):
        self.frame = ttk.Frame(parent, style="Card.TFrame", padding=(12, 12))
        self.frame.pack(fill="both", expand=True, padx=12, pady=12)
        self._build_ui()

    def _build_ui(self):
        fonte_label = font.Font(family="Helvetica", size=10)
        fonte_texto = font.Font(family="Helvetica", size=10)
        fonte_botao = font.Font(family="Helvetica", size=11, weight="bold")

        # Header com logo
        top = ttk.Frame(self.frame, style="Card.TFrame")
        top.pack(fill="x", pady=(0, 8))

        try:
            caminho_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
            if os.path.exists(caminho_logo):
                imagem_logo = Image.open(caminho_logo).resize((64, 64))
                self.logo_tk = ImageTk.PhotoImage(imagem_logo)
                ttk.Label(top, image=self.logo_tk, style="Card.TLabel").pack(side="left", padx=(0, 10))
        except:
            pass

        ttk.Label(top, text="Gerador de Pastas", style="Card.TLabel",
                  font=(default_font[0], 12, "bold")).pack(side="left")

        # Pasta Base
        ttk.Label(self.frame, text="Pasta base (opcional):", style="Card.TLabel").pack(anchor="w", pady=(6, 2))
        frame_base = ttk.Frame(self.frame, style="Card.TFrame")
        frame_base.pack(fill="x", pady=2)

        self.entry_base = tk.Entry(frame_base, bg=INPUT_BG, fg=TEXT,
                                   insertbackground=TEXT, font=fonte_texto, relief="flat")
        self.entry_base.pack(side="left", fill="x", expand=True, ipady=6)

        styled_button(frame_base, text="Procurar", command=self.escolher_pasta).pack(side="left", padx=6)

        # Pasta principal
        ttk.Label(self.frame, text="Nome da pasta principal (opcional):",
                  style="Card.TLabel").pack(anchor="w", pady=(8, 2))
        self.entry_pasta_principal = tk.Entry(self.frame, bg=INPUT_BG, fg=TEXT,
                                              insertbackground=TEXT, font=fonte_texto, relief="flat")
        self.entry_pasta_principal.pack(fill="x", pady=2, ipady=6)

        # Nomes
        ttk.Label(self.frame, text="Nomes (um por linha):", style="Card.TLabel").pack(anchor="w", pady=(8, 2))
        self.text_nomes = tk.Text(self.frame, height=4, bg=INPUT_BG, fg=TEXT,
                                  insertbackground=TEXT, bd=0)
        self.text_nomes.pack(fill="x", pady=2)

        # Subpastas gerais
        ttk.Label(self.frame, text="Subpastas gerais (um por linha):",
                  style="Card.TLabel").pack(anchor="w", pady=(8, 2))
        self.text_subpastas_geral = tk.Text(self.frame, height=3, bg=INPUT_BG, fg=TEXT,
                                            insertbackground=TEXT, bd=0)
        self.text_subpastas_geral.pack(fill="x", pady=2)

        # Subpastas secundárias
        ttk.Label(self.frame, text="Subpastas secundárias (um por linha):",
                  style="Card.TLabel").pack(anchor="w", pady=(8, 2))
        self.text_subpastas_secundarias = tk.Text(self.frame, height=3, bg=INPUT_BG, fg=TEXT,
                                                  insertbackground=TEXT, bd=0)
        self.text_subpastas_secundarias.pack(fill="x", pady=2)

        # Botão Criar
        styled_button(self.frame, text="CRIAR PASTAS", command=self.criar_pastas).pack(
            pady=12, ipadx=6, ipady=6
        )

    def escolher_pasta(self):
        pasta = filedialog.askdirectory(parent=root, title="Selecione a pasta base")
        if pasta:
            self.entry_base.delete(0, tk.END)
            self.entry_base.insert(0, pasta)

    def criar_pastas(self):
        base_path = self.entry_base.get().strip() or filedialog.askdirectory(parent=root, title="Selecione a pasta base")
        if not base_path:
            messagebox.showwarning("Aviso", "Nenhuma pasta base selecionada!")
            return

        nome_principal = self.entry_pasta_principal.get().strip()
        nomes = [n.strip() for n in self.text_nomes.get("1.0", tk.END).split("\n") if n.strip()]
        gerais = [g.strip() for g in self.text_subpastas_geral.get("1.0", tk.END).split("\n") if g.strip()]
        secundarias = [s.strip() for s in self.text_subpastas_secundarias.get("1.0", tk.END).split("\n") if s.strip()]

        if not nomes:
            messagebox.showerror("Erro", "Insira ao menos um nome!")
            return

        pasta_principal = os.path.join(base_path, nome_principal) if nome_principal else base_path

        os.makedirs(pasta_principal, exist_ok=True)

        for n in nomes:
            caminho_nome = os.path.join(pasta_principal, n)
            os.makedirs(caminho_nome, exist_ok=True)

            for g in gerais:
                caminho_g = os.path.join(caminho_nome, g)
                os.makedirs(caminho_g, exist_ok=True)

                for s in secundarias:
                    caminho_s = os.path.join(caminho_g, s)
                    os.makedirs(caminho_s, exist_ok=True)

        messagebox.showinfo("Sucesso", "Pastas criadas com sucesso!")


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    notebook = ttk.Notebook(root, style="TNotebook")
    notebook.pack(fill="both", expand=True, padx=12, pady=12)

    aba1 = ttk.Frame(notebook, style="Card.TFrame")
    notebook.add(aba1, text="Gerador de Pastas")
    GeradorPastas(aba1)

    # Rodapé
    footer = ttk.Frame(root, style="TFrame")
    footer.pack(fill="x", side="bottom")
    ttk.Label(footer, text="• Interface Dark Mode • Profissional • Mantidas todas funcionalidades",
              style="TLabel").pack(pady=6)

    root.mainloop()
